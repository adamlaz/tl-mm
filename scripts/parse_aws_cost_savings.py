#!/usr/bin/env python3
"""Parse Ana's AWS Cost Savings spreadsheet into structured JSON/CSV + ECharts charts.

Reads: AWS Cost Savings.xlsx (Retail, Cake sheets)
Writes:
  - analysis/aws_cost_savings.json
  - analysis/aws_cost_savings.csv
  - analysis/charts/aws_cost_savings_timeline.json
  - analysis/charts/aws_cost_savings_pareto.json
  - analysis/charts/aws_cost_done_vs_remaining.json
"""

import json
import csv
import os
import re
import sys
from collections import defaultdict
from datetime import datetime

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required: pip install openpyxl")

sys.path.insert(0, os.path.dirname(__file__))
import tl_echarts_style as zcs

XLSX_PATH = os.path.join(os.path.dirname(__file__), '..', 'AWS Cost Savings.xlsx')
OUT_DIR = os.path.join(os.path.dirname(__file__), '..', 'analysis')
CHARTS_DIR = os.path.join(OUT_DIR, 'charts')

CATEGORY_PATTERNS = {
    'account_closure': [
        r'closed account', r'deactivated',
    ],
    'env_decommission': [
        r'tst[123]', r'pos2', r'environment decommission', r'shut down all',
        r'sandbox',
    ],
    'k8s_support': [
        r'k8s extended support', r'k8s 1\.\d+',
    ],
    'volume_cleanup': [
        r'volume', r'snapshot', r'ebs', r'gp2 to gp3',
    ],
    'tool_removal': [
        r'sentry', r'prismatic', r'gitlab', r'n8n', r'loki', r'prometheus',
        r'go-?daddy',
    ],
    'architecture_modernization': [
        r'arm64', r'graviton', r'migrated to arm', r'node reduction',
        r'nodegroup',
    ],
    'resource_cleanup': [
        r'nat gateway', r'elastic ip', r'load balancer', r'idle',
        r'stopped', r'unused', r'decommi?ssion(?!.*environment)',
        r'deleted? (?:old|unused)', r'cleanup', r'terminate',
        r'directory service', r'work ?spaces', r'user cleanup',
        r'vpn', r's3.*retired', r'redis', r'memcached',
        r'elasticsearch', r'rds instance',
    ],
}


def classify(task: str) -> str:
    t = task.lower()
    for category, patterns in CATEGORY_PATTERNS.items():
        for p in patterns:
            if re.search(p, t):
                return category
    return 'other'


def parse_currency(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace('$', '').replace(',', '').strip()
    m = re.search(r'[\d.]+', s)
    return float(m.group()) if m else 0.0


def parse_date(val) -> str:
    if val is None:
        return ''
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if not s:
        return ''
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y'):
        try:
            return datetime.strptime(s.split(' ')[0], fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return s


def extract_url(val) -> str:
    if val is None:
        return ''
    s = str(val).strip()
    m = re.search(r'https?://[^\s]+', s)
    return m.group() if m else s


def parse_sheet(ws, domain: str, has_start_date: bool = False):
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    items = []
    for row in rows:
        if has_start_date:
            task, start_date, end_date, loe, monthly, yearly, notes, ticket = (
                row[0], row[1], row[2], row[3], row[4], row[5],
                row[6] if len(row) > 6 else None,
                row[7] if len(row) > 7 else None,
            )
        else:
            task, end_date, loe, monthly, yearly, notes, ticket = (
                row[0], row[1], row[2], row[3], row[4],
                row[5] if len(row) > 5 else None,
                row[6] if len(row) > 6 else None,
            )
            start_date = None

        if not task or not str(task).strip():
            continue

        task_str = str(task).strip()
        yearly_val = parse_currency(yearly)
        monthly_val = parse_currency(monthly)

        if yearly_val == 0 and monthly_val > 0:
            yearly_val = monthly_val * 12
        if monthly_val == 0 and yearly_val > 0:
            monthly_val = round(yearly_val / 12, 2)

        if yearly_val == 0 and monthly_val == 0:
            continue

        items.append({
            'task': task_str,
            'domain': domain,
            'start_date': parse_date(start_date) if has_start_date else '',
            'end_date': parse_date(end_date),
            'monthly_savings': round(monthly_val, 2),
            'yearly_savings': round(yearly_val, 2),
            'notes': str(notes).strip() if notes else '',
            'jira_ticket': extract_url(ticket),
            'category': classify(task_str),
        })
    return items


def build_timeline_chart(items):
    """Stacked bar: cumulative savings by month, Retail vs CAKE."""
    monthly = defaultdict(lambda: {'Retail': 0.0, 'CAKE': 0.0})

    for item in items:
        d = item['end_date']
        if not d:
            d = '2025-09'
        else:
            d = d[:7]
        monthly[d][item['domain']] += item['yearly_savings']

    months = sorted(monthly.keys())
    retail_vals = []
    cake_vals = []
    cumulative_retail = 0.0
    cumulative_cake = 0.0
    for m in months:
        cumulative_retail += monthly[m]['Retail']
        cumulative_cake += monthly[m]['CAKE']
        retail_vals.append(round(cumulative_retail))
        cake_vals.append(round(cumulative_cake))

    labels = [datetime.strptime(m + '-01', '%Y-%m-%d').strftime('%b %Y') if len(m) == 7 else m
              for m in months]

    option = zcs.bar_config(
        labels,
        [
            {'text': 'Retail', 'values': retail_vals, 'backgroundColor': zcs.TL_CATEGORICAL[0]},
            {'text': 'CAKE', 'values': cake_vals, 'backgroundColor': zcs.TL_CATEGORICAL[1]},
        ],
        stacked=True,
        title='Cumulative AWS Cost Savings',
        source='Ana Chambers / Matias Riglos tracking sheet',
        y_title='Annual Savings ($)',
    )
    option['yAxis']['axisLabel']['formatter'] = '{value}'
    option['tooltip']['valueFormatter'] = '${value}'
    return option


def build_pareto_chart(items):
    """Pareto: top initiatives by yearly savings + cumulative %."""
    sorted_items = sorted(items, key=lambda x: -x['yearly_savings'])
    top = sorted_items[:20]
    total = sum(i['yearly_savings'] for i in items)

    labels = [i['task'][:45] + ('...' if len(i['task']) > 45 else '') for i in top]
    values = [round(i['yearly_savings']) for i in top]

    cumulative = []
    running = 0
    for v in values:
        running += v
        cumulative.append(round(running / total * 100, 1))

    option = zcs._base_option('Top Savings Initiatives (Pareto)')
    option['xAxis'] = zcs._cat_axis(labels=labels)
    option['xAxis']['axisLabel']['rotate'] = 35
    option['xAxis']['axisLabel']['fontSize'] = 9
    option['xAxis']['axisLabel']['interval'] = 0
    option['yAxis'] = [
        {**zcs._val_axis(name='Yearly Savings ($)'), 'position': 'left'},
        {**zcs._val_axis(name='Cumulative %'), 'position': 'right', 'max': 100,
         'axisLabel': {**zcs._val_axis()['axisLabel'], 'formatter': '{value}%'}},
    ]
    option['grid']['bottom'] = 100
    option['series'] = [
        {
            'name': 'Yearly Savings',
            'type': 'bar',
            'data': values,
            'itemStyle': {'color': zcs.TL_CATEGORICAL[0]},
            'animationDuration': 600,
            'animationEasing': 'quartOut',
        },
        {
            'name': 'Cumulative %',
            'type': 'line',
            'yAxisIndex': 1,
            'data': cumulative,
            'lineStyle': {'color': zcs.TL_CATEGORICAL[2], 'width': 2},
            'itemStyle': {'color': zcs.TL_CATEGORICAL[2]},
            'symbol': 'circle',
            'symbolSize': 5,
            'animationDuration': 800,
            'animationEasing': 'cubicOut',
            'animationDelay': 300,
        },
    ]
    zcs.add_toolbox(option)
    zcs.add_source_annotation(option, 'Ana Chambers / Matias Riglos tracking sheet')
    return option


def build_done_vs_remaining_chart(items):
    """Horizontal stacked bar: achieved vs estimated remaining by category."""
    achieved_by_cat = defaultdict(float)
    for item in items:
        achieved_by_cat[item['category']] += item['yearly_savings']

    category_labels = {
        'account_closure': 'Account Closures',
        'env_decommission': 'Env Decommissions',
        'k8s_support': 'K8s Support Fees',
        'volume_cleanup': 'Volume/Snapshot Cleanup',
        'tool_removal': 'Tool Removals',
        'architecture_modernization': 'Architecture Modernization',
        'resource_cleanup': 'Resource Cleanup',
        'other': 'Other',
    }

    remaining_estimates = {
        'Monvia Legacy': 311000,
        'Marketplace SaaS': 180000,
        'RDS Right-sizing': 120000,
        'ElastiCache Optimization': 60000,
        'CloudWatch Optimization': 48000,
        'Graviton Migration (EC2)': 80000,
        'Lambda EOL Runtimes': 12000,
    }

    categories = []
    achieved_vals = []
    remaining_vals = []

    for cat_key, cat_label in category_labels.items():
        val = achieved_by_cat.get(cat_key, 0)
        if val > 0:
            categories.append(cat_label)
            achieved_vals.append(round(val))
            remaining_vals.append(0)

    for label, est in remaining_estimates.items():
        categories.append(label)
        achieved_vals.append(0)
        remaining_vals.append(round(est))

    option = zcs._base_option('Achieved Savings vs Remaining Opportunity')
    option['yAxis'] = zcs._cat_axis(labels=categories)
    option['xAxis'] = zcs._val_axis(name='Annual Savings ($)')
    option['grid'] = {'containLabel': True, 'left': 16, 'right': 40, 'top': 56, 'bottom': 48}
    option['series'] = [
        {
            'name': 'Achieved',
            'type': 'bar',
            'stack': 'total',
            'data': achieved_vals,
            'itemStyle': {'color': zcs.TL_STATUS['green']},
            'label': {'show': False},
            'animationDuration': 600,
            'animationEasing': 'quartOut',
        },
        {
            'name': 'Estimated Remaining',
            'type': 'bar',
            'stack': 'total',
            'data': remaining_vals,
            'itemStyle': {'color': zcs.TL_STATUS['amber']},
            'label': {'show': False},
            'animationDuration': 600,
            'animationEasing': 'quartOut',
            'animationDelay': 200,
        },
    ]
    zcs.add_toolbox(option)
    zcs.add_source_annotation(option, 'Ana/Matias tracking sheet + TL AWS scan', date='April 2026')
    return option


def main():
    os.makedirs(CHARTS_DIR, exist_ok=True)

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    all_items = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if sheet_name.strip().lower() in ('sheet1', ''):
            continue

        header = [str(c.value).strip().lower() if c.value else '' for c in ws[1]]
        has_start = 'start date' in header

        domain = 'Retail' if 'retail' in sheet_name.lower() else 'CAKE'
        items = parse_sheet(ws, domain, has_start_date=has_start)
        all_items.extend(items)
        print(f"  {sheet_name}: {len(items)} items, ${sum(i['yearly_savings'] for i in items):,.0f}/yr")

    print(f"\n  Total: {len(all_items)} initiatives, ${sum(i['yearly_savings'] for i in all_items):,.0f}/yr")

    # JSON output
    json_path = os.path.join(OUT_DIR, 'aws_cost_savings.json')
    summary = {
        'generated_at': datetime.now().isoformat(),
        'source': 'AWS Cost Savings.xlsx (Ana Chambers / Matias Riglos)',
        'total_initiatives': len(all_items),
        'total_yearly_savings': round(sum(i['yearly_savings'] for i in all_items), 2),
        'total_monthly_savings': round(sum(i['monthly_savings'] for i in all_items), 2),
        'by_domain': {},
        'by_category': {},
        'initiatives': all_items,
    }
    for item in all_items:
        d = item['domain']
        summary['by_domain'].setdefault(d, {'count': 0, 'yearly': 0})
        summary['by_domain'][d]['count'] += 1
        summary['by_domain'][d]['yearly'] += item['yearly_savings']
        c = item['category']
        summary['by_category'].setdefault(c, {'count': 0, 'yearly': 0})
        summary['by_category'][c]['count'] += 1
        summary['by_category'][c]['yearly'] += item['yearly_savings']

    for d in summary['by_domain'].values():
        d['yearly'] = round(d['yearly'], 2)
    for c in summary['by_category'].values():
        c['yearly'] = round(c['yearly'], 2)

    with open(json_path, 'w') as f:
        json.dump(summary, f, indent=2, ensure_ascii=False)
    print(f"  Wrote {json_path}")

    # CSV output
    csv_path = os.path.join(OUT_DIR, 'aws_cost_savings.csv')
    fieldnames = ['task', 'domain', 'category', 'start_date', 'end_date',
                  'monthly_savings', 'yearly_savings', 'notes', 'jira_ticket']
    with open(csv_path, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for item in all_items:
            writer.writerow({k: item[k] for k in fieldnames})
    print(f"  Wrote {csv_path}")

    # Charts
    timeline = build_timeline_chart(all_items)
    zcs.write_echarts_json(timeline, os.path.join(CHARTS_DIR, 'aws_cost_savings_timeline.json'))
    print(f"  Wrote aws_cost_savings_timeline.json")

    pareto = build_pareto_chart(all_items)
    zcs.write_echarts_json(pareto, os.path.join(CHARTS_DIR, 'aws_cost_savings_pareto.json'))
    print(f"  Wrote aws_cost_savings_pareto.json")

    done_remaining = build_done_vs_remaining_chart(all_items)
    zcs.write_echarts_json(done_remaining, os.path.join(CHARTS_DIR, 'aws_cost_done_vs_remaining.json'))
    print(f"  Wrote aws_cost_done_vs_remaining.json")


if __name__ == '__main__':
    print("=== Parsing AWS Cost Savings spreadsheet ===")
    main()
    print("\nDone.")
